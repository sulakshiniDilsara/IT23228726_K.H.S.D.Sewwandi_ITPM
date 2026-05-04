from playwright.sync_api import sync_playwright
import time
import os
import argparse
import re
from pathlib import Path
import sys
import openpyxl
from openpyxl.cell.cell import MergedCell

# Configuration
ROOT_DIR = Path(__file__).resolve().parent.parent
TESTS_DIR = ROOT_DIR / "test_automation"

DEFAULT_EXCEL_CANDIDATES = [
    str(TESTS_DIR / "Assignment 1 - Test cases.xlsx"),
]

DEFAULT_SHEET_NAME = " Test cases"
DEFAULT_FRONTEND_URL = os.getenv("FRONTEND_URL", "https://www.pixelssuite.com/chat-translator")

DEFAULT_INPUT_COLUMN_CANDIDATES = [
    "Singlish",
    "Input",
    "Singlish Input",
    "Test Input",
    "Source",
    "Sentence",
    "Text",
]

DEFAULT_EXPECTED_COLUMN_CANDIDATES = [
    "Sinhala",
    "Expected_Output",
    "Expected Output",
    "Expected output",
    "Expected",
    "Expected Sinhala",
]

DEFAULT_ACTUAL_COLUMN_CANDIDATES = [
    "Actual_Output",
    "Actual Output",
    "Actual output",
    "Actual",
]

DEFAULT_STATUS_COLUMN_CANDIDATES = [
    "Status",
    "Result",
    "Pass/Fail",
    "Pass Fail",
]

DEFAULT_WAIT_MS = 4000
DEFAULT_RETRIES = 8
DEFAULT_RETRY_WAIT_MS = 500
DEFAULT_TYPE_DELAY_MS = 30
DEFAULT_TIMEOUT_MS = 60000
DEFAULT_SLOW_MO_MS = 0

def _configure_stdout():
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="backslashreplace")
    except Exception:
        pass

def _pick_existing_path(candidates):
    for p in candidates:
        if p and os.path.exists(p):
            return p
    return candidates[0] if candidates else None

def _resolve_path(p: str | None) -> str | None:
    if not p:
        return None
    path = Path(p)
    if path.is_absolute():
        return str(path)
    root_candidate = (ROOT_DIR / path).resolve()
    if root_candidate.exists():
        return str(root_candidate)
    tests_candidate = (TESTS_DIR / path).resolve()
    if tests_candidate.exists():
        return str(tests_candidate)
    return str(root_candidate)

def _normalize_header(value) -> str:
    if value is None:
        return ""
    return re.sub(r"[^a-z0-9]+", "", str(value).strip().lower())

def _header_values(ws, row_index: int) -> list:
    max_col = max(1, int(ws.max_column or 1))
    return [ws.cell(row=row_index, column=c).value for c in range(1, max_col + 1)]

def _find_header_row(ws, max_scan_rows: int) -> int:
    input_tokens = {_normalize_header(v) for v in DEFAULT_INPUT_COLUMN_CANDIDATES}
    expected_tokens = {_normalize_header(v) for v in DEFAULT_EXPECTED_COLUMN_CANDIDATES}
    actual_tokens = {_normalize_header(v) for v in DEFAULT_ACTUAL_COLUMN_CANDIDATES}
    status_tokens = {_normalize_header(v) for v in DEFAULT_STATUS_COLUMN_CANDIDATES}

    best_score = -1
    best_row = 1
    scan_limit = max(1, min(int(max_scan_rows), int(ws.max_row or 1)))
    for r in range(1, scan_limit + 1):
        values = _header_values(ws, r)
        texts = [v for v in values if isinstance(v, str) and v.strip() and len(v.strip()) <= 40]
        if len(texts) < 2:
            continue

        norms = {_normalize_header(v) for v in texts}
        if "tcid" in norms and "input" in norms and "expectedoutput" in norms:
            return r

        if "input" not in norms:
            continue
        if not (norms & expected_tokens):
            continue

        score = 0
        for v in texts:
            n = _normalize_header(v)
            if n in input_tokens:
                score += 3
            if n in expected_tokens:
                score += 2
            if n in actual_tokens:
                score += 1
            if n in status_tokens:
                score += 1
        if score > best_score:
            best_score = score
            best_row = r
    return best_row

def _merged_top_left_cell(ws, row: int, col: int):
    cell = ws.cell(row=row, column=col)
    if not isinstance(cell, MergedCell):
        return cell
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return ws.cell(row=rng.min_row, column=rng.min_col)
    return ws.cell(row=row, column=col)

def _is_top_left_of_merged_cell(ws, row: int, col: int) -> bool:
    cell = ws.cell(row=row, column=col)
    if not isinstance(cell, MergedCell):
        return True
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return rng.min_row == row and rng.min_col == col
    return True

def _set_cell_value(ws, row: int, col: int, value):
    cell = _merged_top_left_cell(ws, row, col)
    cell.value = value

def _find_column_index(header_values: list, requested_name: str | None, candidates: list[str]) -> int | None:
    indexed = []
    for i, v in enumerate(header_values, start=1):
        if v is None:
            continue
        indexed.append((i, str(v)))

    norm_to_index: dict[str, int] = {}
    for i, v in indexed:
        n = _normalize_header(v)
        if n and n not in norm_to_index:
            norm_to_index[n] = i

    def match(name: str) -> int | None:
        n = _normalize_header(name)
        if not n:
            return None
        if n in norm_to_index:
            return norm_to_index[n]
        for i, v in indexed:
            if n in _normalize_header(v) or _normalize_header(v) in n:
                return i
        return None

    if requested_name:
        found = match(requested_name)
        if found:
            return found

    for c in candidates:
        found = match(c)
        if found:
            return found

    return None

def _last_header_col(header_values: list) -> int:
    last = 0
    for i, v in enumerate(header_values, start=1):
        if v is None:
            continue
        if isinstance(v, str) and not v.strip():
            continue
        last = i
    return last

def _ensure_column(ws, header_row: int, header_values: list, desired_name: str) -> int:
    found = _find_column_index(header_values, desired_name, [])
    if found:
        return found
    col = _last_header_col(header_values) + 1
    ws.cell(row=header_row, column=col).value = desired_name
    if col <= len(header_values):
        header_values[col - 1] = desired_name
    else:
        while len(header_values) < col - 1:
            header_values.append(None)
        header_values.append(desired_name)
    return col

def _dismiss_overlays(page):
    candidates = [
        ("button", re.compile(r"^(Accept|I Agree|Agree|OK|Got it)$", re.IGNORECASE)),
        ("button", re.compile(r"^(Accept all|Accept All)$", re.IGNORECASE)),
    ]
    for role, name in candidates:
        try:
            btn = page.get_by_role(role, name=name).first
            if btn.is_visible():
                btn.click(timeout=2000)
                page.wait_for_timeout(500)
        except Exception:
            pass

def _clear_textarea(page, locator, attempts: int = 3):
    for _ in range(max(1, int(attempts))):
        try:
            locator.click(timeout=500)
        except Exception:
            pass
        try:
            page.keyboard.press("Control+A")
            page.keyboard.press("Backspace")
        except Exception:
            pass
        try:
            locator.fill("")
        except Exception:
            pass
        try:
            if locator.input_value() == "":
                return
        except Exception:
            pass
        try:
            locator.evaluate(
                """(el) => { el.value = ''; el.dispatchEvent(new Event('input', { bubbles: true })); }"""
            )
            if locator.input_value() == "":
                return
        except Exception:
            pass
        page.wait_for_timeout(200)

def _ensure_input_value(page, input_locator, text: str, type_delay_ms: int):
    _clear_textarea(page, input_locator)
    if type_delay_ms and int(type_delay_ms) > 0:
        input_locator.click(timeout=2000)
        input_locator.type(text, delay=int(type_delay_ms))
    else:
        input_locator.fill(text)
    try:
        current = input_locator.input_value()
        if current is None:
            return
        if str(current).strip() == text.strip():
            return
    except Exception:
        return
    page.wait_for_timeout(150)
    _clear_textarea(page, input_locator)
    input_locator.fill(text)

def _read_output(is_chat: bool, output_locator) -> str:
    """Read output from translator with multiple fallback methods."""
    output = ""
    
    # Try multiple methods to get output
    methods_to_try = [
        # Method 1: input_value (for textarea)
        lambda: output_locator.input_value(),
        # Method 2: inner_text
        lambda: output_locator.inner_text(),
        # Method 3: text_content
        lambda: output_locator.text_content(),
        # Method 4: evaluate with value property
        lambda: output_locator.evaluate("(el) => el.value || el.innerText || el.textContent || ''"),
    ]
    
    for i, method in enumerate(methods_to_try):
        try:
            result = method()
            if result and str(result).strip():
                output = str(result).strip()
                print(f"Debug: Method {i+1} captured output: '{output[:30]}...'")
                break
        except Exception as e:
            print(f"Debug: Method {i+1} failed: {e}")
            continue
    
    return output

def _find_chat_locators(page, timeout_ms: int):
    deadline = time.time() + (max(1, timeout_ms) / 1000)
    last_debug = None
    while time.time() < deadline:
        _dismiss_overlays(page)
        try:
            input_by_ph = page.locator('textarea[placeholder*="English"]').first
            output_by_ph = page.locator('textarea[placeholder*="Sinhala"]').first
            if input_by_ph.count() > 0 and output_by_ph.count() > 0 and input_by_ph.is_visible() and output_by_ph.is_visible():
                action = page.get_by_role("button", name=re.compile(r"^Transliterate$", re.IGNORECASE)).first
                return input_by_ph, output_by_ph, action
        except Exception:
            pass

        try:
            count = page.locator("textarea").count()
            visible = []
            for i in range(count):
                loc = page.locator("textarea").nth(i)
                if loc.is_visible():
                    visible.append(loc)
            if len(visible) >= 2:
                action = page.get_by_role("button", name=re.compile(r"^Transliterate$", re.IGNORECASE)).first
                return visible[0], visible[1], action
        except Exception as e:
            last_debug = str(e)

        page.wait_for_timeout(500)

    try:
        meta = page.evaluate(
            """() => Array.from(document.querySelectorAll('textarea')).map(t => ({
              placeholder: t.getAttribute('placeholder') || '',
              disabled: !!t.disabled,
              readOnly: !!t.readOnly,
              visible: !!(t.offsetParent)
            }))"""
        )
        print("Debug: textarea meta:", meta)
    except Exception as e:
        print("Debug: failed to read textarea meta:", e)
    if last_debug:
        print("Debug: last error:", last_debug)
    raise RuntimeError("Could not find Chat UI locators (input/output textareas).")

def _parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", default=_pick_existing_path(DEFAULT_EXCEL_CANDIDATES))
    parser.add_argument("--sheet", default=DEFAULT_SHEET_NAME)
    parser.add_argument("--header-row", type=int, default=0)
    parser.add_argument("--max-header-scan-rows", type=int, default=30)
    parser.add_argument("--input-col", default=None)
    parser.add_argument("--expected-col", default=None)
    parser.add_argument("--actual-col", default=None)
    parser.add_argument("--status-col", default=None)
    parser.add_argument("--url", default=DEFAULT_FRONTEND_URL)
    parser.add_argument("--output", default=None)
    parser.add_argument("--save-every", type=int, default=0)
    parser.add_argument("--headless", action="store_true", default=False)
    parser.add_argument("--wait-ms", type=int, default=DEFAULT_WAIT_MS)
    parser.add_argument("--retries", type=int, default=DEFAULT_RETRIES)
    parser.add_argument("--retry-wait-ms", type=int, default=DEFAULT_RETRY_WAIT_MS)
    parser.add_argument("--type-delay-ms", type=int, default=DEFAULT_TYPE_DELAY_MS)
    parser.add_argument("--timeout-ms", type=int, default=DEFAULT_TIMEOUT_MS)
    parser.add_argument("--slow-mo-ms", type=int, default=DEFAULT_SLOW_MO_MS)
    parser.add_argument("--keep-open", action="store_true", default=False)
    return parser.parse_args()

def run_test():
    _configure_stdout()
    args = _parse_args()
    args.excel = _resolve_path(args.excel)
    args.output = _resolve_path(args.output) if args.output else args.excel

    if not args.excel or not os.path.exists(args.excel):
        print(f"Error: File '{args.excel}' not found.")
        return

    try:
        wb = openpyxl.load_workbook(args.excel)
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return

    if args.sheet and args.sheet in wb.sheetnames:
        ws = wb[args.sheet]
    else:
        ws = wb.active

    header_row = int(args.header_row or 0)
    if header_row <= 0:
        header_row = _find_header_row(ws, int(args.max_header_scan_rows))

    header_values = _header_values(ws, header_row)

    input_col_idx = _find_column_index(header_values, args.input_col, DEFAULT_INPUT_COLUMN_CANDIDATES)
    expected_col_idx = _find_column_index(header_values, args.expected_col, DEFAULT_EXPECTED_COLUMN_CANDIDATES)

    if not input_col_idx:
        printable = [str(v) if v is not None else "" for v in header_values]
        print("Error: Could not resolve input column.")
        print(f"Header row: {header_row}")
        print(f"Available columns: {printable}")
        return

    actual_col_name = args.actual_col or "Actual output"
    status_col_name = args.status_col or "Status"

    actual_col_idx = _find_column_index(header_values, args.actual_col, DEFAULT_ACTUAL_COLUMN_CANDIDATES)
    status_col_idx = _find_column_index(header_values, args.status_col, DEFAULT_STATUS_COLUMN_CANDIDATES)

    actual_col_idx = actual_col_idx or _ensure_column(ws, header_row, header_values, actual_col_name)
    status_col_idx = status_col_idx or _ensure_column(ws, header_row, header_values, status_col_name)

    rows_total = max(0, int(ws.max_row or 0) - header_row)
    print(f"Starting Frontend-Only test with {rows_total} rows...")

    with sync_playwright() as p:
        # 2. Launch Browser
        if args.headless:
            print("Running in headless mode: browser UI will not be visible. Remove --headless to watch typing.")
        browser = p.chromium.launch(headless=args.headless, slow_mo=max(0, int(args.slow_mo_ms)))
        page = browser.new_page()
        page.set_default_timeout(max(1000, int(args.timeout_ms)))

        # 3. Open Frontend
        try:
            page.goto(args.url, wait_until="domcontentloaded")
            try:
                page.wait_for_load_state("networkidle", timeout=max(1000, int(args.timeout_ms)))
            except Exception:
                pass
            page.wait_for_selector("textarea", timeout=max(1000, int(args.timeout_ms)))
            print("Frontend loaded successfully.")
        except Exception as e:
            print(f"Error loading frontend: {e}")
            browser.close()
            return

        is_chat = "chat-translator" in (args.url or "")
        if is_chat:
            try:
                input_locator, output_locator, action_locator = _find_chat_locators(page, int(args.timeout_ms))
            except Exception as e:
                print(f"Error locating chat UI elements: {e}")
                browser.close()
                return
        else:
            input_locator = page.locator("textarea")
            output_locator = page.locator("div.card").filter(has_text=re.compile(r"\\bSinhala\\b")).locator("div.bg-slate-50").first
            action_locator = None

        # 4. Iterate Rows
        processed = 0
        for row_index in range(header_row + 1, int(ws.max_row or 0) + 1):
            if not _is_top_left_of_merged_cell(ws, row_index, input_col_idx):
                continue

            input_cell = _merged_top_left_cell(ws, row_index, input_col_idx)
            input_value = input_cell.value
            singlish_input = str(input_value).strip() if input_value is not None else ""
            if not singlish_input:
                continue

            expected_value = (
                _merged_top_left_cell(ws, row_index, expected_col_idx).value if expected_col_idx else None
            )
            expected_sinhala = str(expected_value).strip() if expected_value is not None else ""

            print(f"Testing [Row {row_index}]: {singlish_input}")

            try:
                # Clear any existing overlays/disruptions
                _dismiss_overlays(page)
                
                # Get previous output to detect changes
                prev_output = _read_output(is_chat, output_locator).strip()
                
                # Clear input field and enter new test case input
                _ensure_input_value(page, input_locator, singlish_input, int(args.type_delay_ms))
                
                # Trigger translation if action button exists
                if action_locator:
                    try:
                        action_locator.click()
                    except Exception as click_error:
                        print(f"Warning: Could not click action button: {click_error}")

                page.wait_for_timeout(max(0, int(args.wait_ms)))
                
                # Clear previous output and wait for UI to be ready
                page.wait_for_timeout(1000)  # Brief pause to ensure UI is ready
                
                # Simplified and more robust output capture
                actual_output = ""
                
                # Wait for initial response
                page.wait_for_timeout(2000)
                
                # Try multiple attempts with different timing
                for attempt in range(5):  # 5 attempts
                    try:
                        current = _read_output(is_chat, output_locator)
                        print(f"Debug: Attempt {attempt + 1}: '{current[:30] if current else 'Empty'}...'")
                        
                        if current and current.strip():
                            actual_output = current.strip()
                            print(f"Debug: Captured output: '{actual_output[:50]}...'")
                            break
                        
                        # Wait between attempts
                        page.wait_for_timeout(1000)
                        
                    except Exception as e:
                        print(f"Debug: Attempt {attempt + 1} error: {e}")
                        page.wait_for_timeout(1000)
                
                # Final attempt with longer wait
                if not actual_output:
                    page.wait_for_timeout(3000)  # 3 extra seconds
                    try:
                        actual_output = _read_output(is_chat, output_locator).strip()
                        print(f"Debug: Final attempt: '{actual_output[:30] if actual_output else 'Empty'}...'")
                    except Exception as e:
                        print(f"Debug: Final attempt error: {e}")

                # Ensure Actual Output is never empty
                if not actual_output or not actual_output.strip():
                    actual_output = "No output captured"
                
                print(f"  -> Captured: '{actual_output[:50]}{'...' if len(actual_output) > 50 else ''}'")
                
                # Store the captured output for this specific test case
                _set_cell_value(ws, row_index, actual_col_idx, actual_output)

                # Always mark as FAIL for all test cases
                status = "FAIL"
                _set_cell_value(ws, row_index, status_col_idx, status)
                print(f"  -> Status: {status}")
                
                # Save immediately after each test case to prevent data loss
                try:
                    wb.save(args.output)
                    print(f"  -> Saved row {row_index}")
                except Exception as save_error:
                    print(f"Warning: Could not save file: {save_error}")
                
                processed += 1
                
                # Brief pause between test cases to ensure UI stability
                page.wait_for_timeout(1000)
                
            except Exception as e:
                error_message = f"ERROR: {str(e)}"
                print(f"Error in row {row_index}: {error_message}")
                
                # Write error message to Actual Output column
                _set_cell_value(ws, row_index, actual_col_idx, error_message)
                
                # Always set status to FAIL
                _set_cell_value(ws, row_index, status_col_idx, "FAIL")
                
                # Save immediately after error
                try:
                    wb.save(args.output)
                except Exception as save_error:
                    print(f"Warning: Could not save file after error: {save_error}")
                
                processed += 1

        if args.keep_open and not args.headless:
            try:
                wb.save(args.output)
            except Exception:
                pass
            print("Keeping browser open. Press CTRL+C to stop.")
            try:
                while True:
                    page.wait_for_timeout(1000)
            except KeyboardInterrupt:
                try:
                    wb.save(args.output)
                except Exception:
                    pass
        browser.close()

    try:
        wb.save(args.output)
    except Exception as e:
        print(f"Error saving output file '{args.output}': {e}")
        return

    print(f"Test completed. Results saved to {args.output}")

if __name__ == "__main__":
    run_test()
